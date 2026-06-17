"""
Geração de planilhas XLSX para os exports das abas de Consultas SIGCON.

Usa openpyxl puro — NUNCA importar pandas/numpy aqui. O processo web roda
com múltiplos workers e reimportar essas libs nesse caminho reproduz o
erro de inicialização do OpenBLAS visto em produção.
"""

from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse

_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_LARGURA_MIN = 10
_LARGURA_MAX = 60


def exportar_xlsx(linhas, colunas, nome_arquivo: str) -> HttpResponse:
    """
    Monta um .xlsx com `colunas` como cabeçalho (negrito) e `linhas` como
    corpo — cada linha é uma sequência de valores na mesma ordem de `colunas`.

    `nome_arquivo` é o nome base (sem extensão e sem timestamp); o timestamp
    de geração é acrescentado automaticamente ao nome do arquivo baixado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    ws.append(list(colunas))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for linha in linhas:
        ws.append([_valor_serializavel(v) for v in linha])

    for idx, label in enumerate(colunas, start=1):
        largura = max(_LARGURA_MIN, min(_LARGURA_MAX, len(str(label)) + 4))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = largura

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    response = HttpResponse(content_type=_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}_{timestamp}.xlsx"'
    wb.save(response)
    return response


def _valor_serializavel(valor):
    """openpyxl grava str/int/float/bool/date/datetime/None direto; Decimal precisa virar float."""
    if valor is None or isinstance(valor, (str, int, float, bool, date, datetime)):
        return valor
    return float(valor)
